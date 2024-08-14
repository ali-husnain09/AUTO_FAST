import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from fake_useragent import UserAgent
from fuzzywuzzy import process
import time
import openpyxl


class DataProcessor:
    def __init__(self, file_path, last_row_file):
        self.file_path = file_path
        self.last_row_file = last_row_file
        self.last_row_number = self.load_last_row_number()

    def get_next_row_data(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        # Assuming the columns are named Name, Address, City, State, and Phone Number
        name = sheet.cell(row=self.last_row_number + 2, column=1).value
        address = sheet.cell(row=self.last_row_number + 2, column=5).value
        city = sheet.cell(row=self.last_row_number + 2, column=6).value
        state = sheet.cell(row=self.last_row_number + 2, column=7).value
        phone_number = sheet.cell(row=self.last_row_number + 2, column=8).value
        email_address = sheet.cell(row=self.last_row_number + 2, column=9).value

        return name, address, city, state, phone_number

    def save_last_row_number(self, last_row_number):
        with open(self.last_row_file, "w") as file:
            file.write(str(last_row_number))

    def load_last_row_number(self):
        try:
            with open(self.last_row_file, "r") as file:
                return int(file.read())
        except FileNotFoundError:
            return 0


# Usage example
file_path = "path_to_your_excel_file.xlsx"
last_row_file = "last_row_number.txt"

processor = DataProcessor(file_path, last_row_file)
name, address, city, state, phone_number = processor.get_next_row_data()

# Process the data (e.g., interact with the TruePeopleSearch website)

# After processing, save the last row number
processor.save_last_row_number(processor.last_row_number + 1)


class PhoneSearchBot:
    def __init__(self, file_path, last_row_file):
        self.file_path = file_path
        self.last_row_file = last_row_file
        self.last_row_number = self.load_last_row_number()
        self.driver = self.setup_driver()

    def setup_driver(self):
        service = Service(executable_path="Apps/chromedriver.exe")
        chrome_options = Options()
        chrome_options.add_extension("Apps/cblock.crx")
        chrome_options.add_extension("Apps/cappass.crx")
        chrome_options.headless = False
        chrome_options.add_argument("--blink-settings=imagesEnabled=false")
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument("--window-size=1920,1080")
        user_agent = UserAgent().random
        chrome_options.add_argument(f"user-agent={user_agent}")

        driver = webdriver.Chrome(service=service, options=chrome_options)
        return driver

    def load_last_row_number(self):
        try:
            with open(self.last_row_file, "r") as file:
                last_row_number = int(file.read())
            return last_row_number
        except FileNotFoundError:
            return 0

    def get_next_row_data(self):
        # Implement logic to read the next row data from the Excel file
        # Return the name, address, city, and state as a tuple
        pass

    def save_last_row_number(self, last_row_number):
        with open(self.last_row_file, "w") as file:
            file.write(str(last_row_number))

    def search_phone_numbers(self):
        while True:
            name, address, city, state, _ = self.get_next_row_data()

            # Address search navigation and input
            input_element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//a[contains(text(),'Address Search')]")
                )
            )
            input_element.click()

            # Address input and search button click
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//input[@id='Address']"))
            ).send_keys(address)
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//input[@id='CityStateZip']")
                )
            ).send_keys(city + "," + state)
            WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[@id='btnSearch']"))
            ).click()

            # Handle search result conditions and extract phone numbers
            # ...

            # Update last row number and save to file
            self.last_row_number += 1
            self.save_last_row_number(self.last_row_number)

            # Check if there are no more rows to process
            if not self.get_next_row_data()[0]:
                break

    def run(self):
        self.driver.get("https://www.fastpeoplesearch.com/")
        WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.NAME, "q"))
        )
        self.search_phone_numbers()
        self.driver.quit()


# Usage
bot = PhoneSearchBot("path/to/excel_file.xlsx", "path/to/last_row_number_file.txt")
bot.run()
