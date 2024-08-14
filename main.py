import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from thefuzz import process
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
# from fake_useragent import UserAgent

from multiprocessing import Process
# import exel_path
import os

# Provide the path to the Text Mode extension CRX file
print("AUTOTRUE: the hunt for valuable bytes.")
print(".......................[DETAILS]............................")


def process_file(file_path, last_row_file):
    # Initialize WebDriver
    service = Service(executable_path="Apps/chromedriver.exe")
    chrome_options = Options()
    chrome_options.binary_location = (
        r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
    )

    # Add Chrome extension
    chrome_options.add_extension("Apps/cblock.crx")
    chrome_options.add_extension("Apps/cappass.crx")

    chrome_options.headless = False
    chrome_options.add_argument("--blink-settings=imagesEnabled=false")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--window-size=1920,1080")
    # user_agent = UserAgent().random
    # chrome_options.add_argument(f"user-agent={user_agent}")

    # Specify user profile directory

    # Initialize Chrome webdriver with specified options
    driver = webdriver.Chrome(service=service, options=chrome_options)

    # Open TruePeopleSearch website
    driver.get("https://www.truepeoplesearch.com/")

    time.sleep(2)

    last_row_number = load_last_row_number(last_row_file)

    while True:
        name, address, city, state, _ = get_next_row_data(file_path, last_row_number)

        # Go to Address Search
        input_element = driver.find_element(
            By.XPATH, "//span[contains(text(),'Address Search')]"
        )
        input_element.click()
        time.sleep(2)

        # Input Address, City, and State information
        input_address = driver.find_element(By.XPATH, "//input[@id='id-d-addr']")
        input_address.send_keys(address)
        time.sleep(2)

        input_address = driver.find_element(By.XPATH, "//input[@id='id-d-loc-addr']")
        input_address.send_keys(city + "," + state)
        time.sleep(2)

        # Click on search button
        driver.implicitly_wait(2)  # seconds
        input_address = driver.find_element(
            By.XPATH, "//button[@id='btnSubmit-d-addr']"
        )
        input_address.send_keys(Keys.ENTER)

        time.sleep(3)

        if (
            "We could not find any records associated with that address."
            in driver.page_source
        ):
            print("No INDEX")
            no_index(file_path, last_row_number, "NOREC")
            driver.back()
            # last_row_number += 1
            continue

        if "Please check the box below." in driver.page_source:
            print("trying to bypass..")
            time.sleep(15)

        else:
            print("Change your VPN")
# //////////////////////////////////////////////////////////////
        try:
            # Extracting names from search results
            output_names = driver.find_elements(By.CLASS_NAME, "h4")

            # Extracting data from the website
            time.sleep(2)

            # Perform fuzzy matching
            texts = [element.text for element in output_names]
            input_name = name

            # Initialize variables to store the best match and its score
            best_matches = []
            best_scores = []

            # Iterate over each extracted text and perform fuzzy matching separately
            for text in texts:
                match, score = process.extractOne(
                    input_name, text.split("\n")
                )  # Assuming text is separated by newline characters
                if score >= 50:  # Check if the score is greater than or equal to 80
                    best_matches.append(match)
                    best_scores.append(score)

            # Sort the matches based on scores
            sorted_indices = sorted(
                range(len(best_scores)), key=lambda k: best_scores[k], reverse=True
            )

            # Iterate over the top three best matches
            for i in range(min(3, len(sorted_indices))):
                best_match = best_matches[sorted_indices[i]]
                best_score = best_scores[sorted_indices[i]]

                # Print the top three best matches
                print(f"Top {i + 1} match: {best_match}, Score: {best_score}")

                # Click on the best matching result
                input_address = driver.find_element(
                    By.XPATH, f"//*[contains(text(), '{best_match}')]"
                )
                input_address.click()

                time.sleep(1)

                try:
                    # Look for the element containing the phone numbers
                    phone_numbers_element = driver.find_element(
                        By.XPATH, '//*[@id="personDetails"]/div[9]'
                    )  # Adjust the class name if needed

                    # Extract text from the element
                    phone_numbers_text = phone_numbers_element.text

                    # Save the phone numbers if found
                    if "Phone Numbers" in phone_numbers_text:
                        phone_numbers = phone_numbers_text
                        save_phone_number(file_path, last_row_number, phone_numbers)
                        print("🤩 Phone numbers found for:", best_match)
                        break  # Break the loop if phone numbers are found
                    else:
                        print("😔 Phone numbers not found for:", best_match)
                        # Navigate back to the homepage if phone numbers are not found
                        driver.back()
                        print("🔙 Backing...")
                except NoSuchElementException:
                    print("Phone numbers not found for:", best_match)
                    # Navigate back to the homepage if phone numbers are not found
                    driver.back()
                    print("🔙 Backing...")

            else:
                print("🙎🏻‍♂️ No More match found with a score of 50 or above")

        except Exception as e:
            print(f"🚩 An error occurred: {e}")

        finally:
            # Navigate back to the homepage
            print("........................[DETAILS]............................")
            driver.get("https://www.truepeoplesearch.com/")
            # Increment last row number for the next run
            last_row_number += 1
            save_last_row_number(last_row_file, last_row_number)

            # Check if there are more rows to process, if not break out of the loop
            if not get_next_row_data(file_path, last_row_number)[0]:
                break

    # Close the WebDriver
    driver.quit()


def get_next_row_data(file_path, last_row_number):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Assuming the columns are named Name, Address, City, State, and Phone Number
    name = sheet.cell(row=last_row_number + 2, column=1).value
    address = sheet.cell(row=last_row_number + 2, column=5).value
    city = sheet.cell(row=last_row_number + 2, column=6).value
    state = sheet.cell(row=last_row_number + 2, column=7).value
    phone_number = sheet.cell(row=last_row_number + 2, column=8).value
    email_address = sheet.cell(row=last_row_number + 2, column=9).value

    return name, address, city, state, phone_number


def save_last_row_number(last_row_file, last_row_number):
    with open(last_row_file, "w") as file:
        file.write(str(last_row_number))


def load_last_row_number(last_row_file):
    try:
        with open(last_row_file, "r") as file:
            return int(file.read())
    except FileNotFoundError:
        return 0


def save_phone_number(file_path, row_number, phone_numbers):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Remove commas from phone numbers and concatenate them into one string
    phone_numbers_str = "".join([phone.replace(",", "") for phone in phone_numbers])

    # Assuming the phone number column is column H (column index 8)
    sheet.cell(row=row_number + 2, column=8).value = phone_numbers_str

    wb.save(file_path)

    # Print the phone numbers in the terminal
    print("Phone Numbers Saved")


def no_index(file_path, row_number, no_Index):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    # Assuming the phone number column is column H (column index 8)
    sheet.cell(row=row_number + 2, column=8).value = no_Index

    wb.save(file_path)
    print("NOTE:", no_Index)


if __name__ == "__main__":
    # Define file paths
    file_paths = [
        (exel_path.P1, "last_row_number1.txt"),
        (exel_path.P2, "last_row_number2.txt"),
    ]

    # Create processes for each file
    processes = []
    for file_path, last_row_file in file_paths:
        p = Process(target=process_file, args=(file_path, last_row_file))
        p.start()
        processes.append(p)

    # Wait for all processes to finish
    for p in processes:
        p.join()
