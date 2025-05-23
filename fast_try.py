from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException

# from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from thefuzz import fuzz
import json
import time
import random
from name_check import name_validations as nValidate
from address_check import address_validations as aValidate
import openpyxl
from colorama import Fore
import undetected_chromedriver as uc
import os

class Iterations(nValidate, aValidate):
    def __init__(self, file_path, last_row_file, variables):
        self.var = variables
        self.file_path = file_path
        self.last_row_file = last_row_file
        self.last_row_number = self.load_last_row_number()
        self.driver = self.setup_driver()
        self.noRecordsFound = (
            "We could not find any records associated with that address."
        )

    def setup_driver(self):
        chrome_options = Options()
        chrome_options.add_argument(
            "--user-data-dir=C:\\Users\\Kali\\AppData\\Local\\Google\\Chrome\\User Data\\Default"
        )
        chrome_options.page_load_strategy = "eager"
        # driver = webdriver.Chrome(
        #     service=ChromeService(ChromeDriverManager().install()),
        #     options=chrome_options,
        # )
        driver = uc.Chrome(options=chrome_options)
        return driver

    def sleeper(self):
        time.sleep(
            float(
                "0."
                + random.choice(timers[0:4])
                + random.choice(timers[0:7])
                + random.choice(timers[0:9])
            )
        )

    def load_last_row_number(self):
        try:
            with open(self.last_row_file, "r") as file:
                last_row_number = int(file.read())
            return last_row_number
        except FileNotFoundError:
            return 0

    def get_next_row_data(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        # Assuming the columns are named Name, Address, City, State, and Phone Number
        user_name = sheet.cell(row=self.last_row_number + 2, column=1).value
        formal_address = sheet.cell(row=self.last_row_number + 2, column=5).value
        city = sheet.cell(row=self.last_row_number + 2, column=6).value
        state = sheet.cell(row=self.last_row_number + 2, column=7).value

        return user_name, formal_address, city, state

    def save_last_row_number(self, last_row_number):
        with open(self.last_row_file, "w") as file:
            file.write(str(last_row_number))

    def read_email_DATA(self):  # READ DATA FROM THE EMAIL TEXT FILE
        emails = []
        with open("emails.txt", "r") as file:
            for line in file:
                emails.append(line)

        return emails  # RETURN EMAIL LIST

    def read_data(self):  # READ DATA FROM THE PHONENUMBERS TEXT FILE
        phone_numbers = []
        with open("phone_numbers.txt", "r") as file:
            for line in file:
                phone_numbers.append(line)

        return phone_numbers  # RETURN PHONENUMBERS LIST

    def save_Value(self, value):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active
        try:
            # Assuming the phone number column is column H (column index 8)
            sheet.cell(row=self.last_row_number + 2, column=8).value = value

        except ValueError:
            # Remove commas from phone numbers and concatenate them into one string
            value = "".join([phone.replace(",", "\n") for phone in value])
            sheet.cell(row=self.last_row_number + 2, column=8).value = value

        # Print the phone numbers in the terminal
        print(Fore.GREEN + f"Provided Value Is Saved is {self.last_row_number + 2}")
        wb.save(self.file_path)

    def save_EMAIL(self, email_val):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active
        try:
            # Assuming the emails column is column I (column index 9)
            sheet.cell(row=self.last_row_number + 2, column=9).value = email_val

        except ValueError:
            # Remove commas from emails list and concatenate them into one string
            email_val_str = "".join([email.replace(",", "\n") for email in email_val])
            sheet.cell(row=self.last_row_number + 2, column=9).value = email_val_str

        # Print the emails_print in the terminal
        print(Fore.GREEN + f"Email Saved is {self.last_row_number + 2}")
        wb.save(self.file_path)

    def check_name(self, user_name):
        return nValidate.__checkValid__(self, user_name)

    def check_address(self, formal_address, city, state):
        return aValidate.__checkValid__(self, formal_address, city, state)

    def search_bests(self):
        self.enteringDetails = True
        while self.enteringDetails:
            user_name, formal_address, city, state = self.get_next_row_data()
            if not user_name:
                print(Fore.LIGHTYELLOW_EX + "No More Rows To Process")
                break
            if self.check_name(user_name):
                print(Fore.YELLOW + "Company Found.")
                # integrate here the value
                self.save_Value("Its Company")
            else:
                if self.check_address(formal_address, city, state):
                    print(Fore.YELLOW + "Address Matched.")
                    try:
                        # phone numbers received to through address
                        self.save_Value(self.read_data())
                        self.save_EMAIL(self.read_email_DATA())
                        print(Fore.BLUE + "Phone Numbers Received")
                        print(Fore.BLUE + "Emails Received")
                    except FileNotFoundError:
                        print(Back.LIGHTRED_EX + Back.BLACK + "DETAILS ARE EMPTY")
                        print(
                            Fore.LIGHTYELLOW_EX
                            + "\n============================================================"
                        )
                    except:
                        print(Back.LIGHTRED_EX + Back.BLACK + "DETAILS ARE EMPTY")
                        print(
                            Fore.LIGHTYELLOW_EX
                            + "\n============================================================"
                        )
                else:

                    try:
                        time.sleep(2)
                        input_element = WebDriverWait(self.driver, 5).until(
                            EC.element_to_be_clickable((By.ID, self.var[0]))
                        )
                        input_element.click()
                        address_bar = self.driver.find_element(By.ID, self.var[1])
                        # address_bar.send_keys(str(formal_address))
                        for i in str(formal_address):
                            address_bar.send_keys(i)
                        # self.sleeper()
                        time.sleep(0.4)
                        city_bar = self.driver.find_element(By.ID, self.var[2])
                        # city_bar.send_keys(str(city + "," + state))
                        for i in str(city + ", " + state):
                            city_bar.send_keys(i)

                        # self.sleeper()
                        time.sleep(0.4)
                        self.driver.find_element(By.XPATH, self.var[3]).click()

                        # Wait for search results to load
                        time.sleep(2)
                        WebDriverWait(self.driver, 20).until(
                            EC.presence_of_all_elements_located(
                                (By.CLASS_NAME, "larger")
                            )
                        )

                        # Extracting names from the search results
                        titles = self.driver.find_elements(By.CLASS_NAME, "larger")
                        output_names = [names.text for names in titles]

                        # Check if output_names is available
                        print(Fore.WHITE + f"Output Names: {output_names}")

                        if not output_names:
                            print(
                                Back.LIGHTRED_EX
                                + Fore.BLACK
                                + "\nNO SEARCH RESULTS ARE AVAILABLE"
                            )
                            if os.path.exists("phone_numbers.txt"):
                                os.remove("phone_numbers.txt")
                            if os.path.exists("emails.txt"):
                                os.remove("emails.txt")
                          
                        # Save the output names to a file
                        with open("output_names.txt", "w") as output:
                            for name in output_names:
                                output.write(f"{name}\n")

                        best_matches = []

                        # Iterate over the names to find best matches
                        for index in output_names:
                            # Print the current comparison
                            print(
                                Fore.WHITE
                                + f"Comparing '{user_name.lower()}' with '{index.lower()}'"
                            )
                            ratio = fuzz.ratio(user_name.lower(), index.lower())
                            print(
                                Fore.GREEN + f"Match Ratio: {ratio}"
                            )  # Print the calculated ratio
                            if ratio >= 40:  # adjust the ratio to 50
                                best_matches.append((index, ratio))

                        # Check if best_matches are available
                        print(Fore.LIGHTYELLOW_EX + f"Best Matches: {best_matches}")

                        # Sort matches by ratio in descending order from highest to lowest
                        best_matches.sort(key=lambda x: x[1], reverse=True)

                        # Save the best matches to a JSON file
                        with open("best_matches.json", "w") as outfile:
                            json.dump(best_matches, outfile)

                        # Print the best matches
                        print(
                            Fore.RED
                            + f"Best matches found and saved to JSON: {best_matches}"
                        )

                        # Click on each best match and perform an action
                        time.sleep(0.5)
                        for i, (name, ratio) in enumerate(best_matches):
                            print(
                                Fore.WHITE + f"Top {i+1} match: {name}, Score: {ratio}"
                            )
                            element = WebDriverWait(self.driver, 20).until(
                                EC.element_to_be_clickable(
                                    (By.XPATH, f"//*[contains(text(), '{name}')]")
                                )
                            )
                            element.click()
                            # self.driver.implicitly_wait(20)  # Wait for the page to load
                            time.sleep(0.5)
                            try:
                                # Look for the element containing the phone numbers
                                phone_numbers_element = WebDriverWait(
                                    self.driver, 3
                                ).until(
                                    EC.presence_of_element_located(
                                        (By.CLASS_NAME, self.var[4])
                                    )
                                )
                                print(Fore.BLUE + f"Phone numbers are found")
                                self.save_Value(phone_numbers_element.text)
                                with open("phone_numbers.txt", "w") as f:
                                    f.write(phone_numbers_element.text)
                                try:
                                    email_element = WebDriverWait(self.driver, 2).until(
                                        EC.presence_of_element_located(
                                            (By.XPATH, self.var[5])
                                        )
                                    )
                                    print(Fore.BLUE + f"Emails are found")
                                    self.save_EMAIL(email_element.text)
                                    with open("emails.txt", "w") as f:
                                        f.write(email_element.text)
                                except:
                                    print(Fore.RED + f"Email not found for: {name}")
                                break  # end the loop for the next data element

                            # //changes here to make a backup
                            except TimeoutException as e:
                                print(Fore.RED + f"Phone numbers not found for: {name}")
                                self.driver.back()  # Go back to search results

                            if i == len(best_matches) - 1:
                                print(
                                    Fore.RED
                                    + "No phone numbers found for all best matches."
                                )
                                break
                    # timeout error for each element
                    except TimeoutException:
                        print(Fore.RED + f"Timeout occurred")
                        self.save_Value("Timeout occurred")
                    # if there are no elements
                    except NoSuchElementException as e:
                        print(Fore.RED + f"Element not found {e}")
                        self.save_Value("NoSuchElement")
                    # others any errors and warnings
                    except Exception as e:
                        print(Fore.RED + f"An error occurred: {e}")
                        self.save_Value("Erros")

            # Do increment for each iteration
            self.last_row_number += 1
            self.save_last_row_number(self.last_row_number)
            self.driver.get(
                "https://www.fastpeoplesearch.com/"
            )  # Go back to search results

        # Finally, Save the file

    def run(self):
        self.driver.get("https://www.fastpeoplesearch.com/")
        self.search_bests()
        self.driver.quit()


timers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]
variables = [
    "search-nav-link-address",
    "search-address-1",
    "search-address-2",
    '//*[@id="form-search-address"]/div[3]/button[2]',
    "detail-box-phone",
    '//*[@id="email_section"]/div/div',
]
file_path = "mytest2.xlsx"  # file path for search results
last_row_file = "last_row_number.txt"  # file path for last row for each iteration
if __name__ == "__main__":
    iterations = Iterations(file_path, last_row_file, variables)
    iterations.run()
    
